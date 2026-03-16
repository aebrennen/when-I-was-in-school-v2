"""Extract statewide Free/Reduced-Price Lunch percentages from KDE Qualifying Data files.

Reads: data/raw/kde_frl/*FinalQualifyingData.xlsx
Produces: data/processed/kde_frl_statewide.json

Each file contains school-level data with columns:
  Total Enrollment, Number Free, Number Reduced, Number Paid,
  Percent Free, Percent Reduced, Percent Free and Reduced

We aggregate to statewide totals per school year.
"""

import json
import os
import re
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
RAW_DIR = os.path.join(DATA_DIR, 'raw', 'kde_frl')
PROCESSED_DIR = os.path.join(DATA_DIR, 'processed')


def extract_file(filepath):
    """Extract statewide FRL totals from a single KDE qualifying data file."""
    filename = os.path.basename(filepath)

    # Parse school year from filename: "2024-2025FinalQualifyingData.xlsx"
    m = re.match(r'(\d{4})-(\d{4})', filename)
    if not m:
        return None
    start_year = int(m.group(1))
    end_year = int(m.group(2))
    school_year = f"{start_year}-{end_year}"

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row — look for "Total Enrollment" or "Total\nEnrollment"
    header_row = None
    col_enrollment = None
    col_free = None
    col_reduced = None
    col_paid = None

    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').lower().replace('\n', ' ').replace('_x000d_', ' ')
            if 'total' in val and 'enrollment' in val:
                header_row = r
                col_enrollment = c
                break
        if header_row:
            break

    if header_row is None:
        print(f"    No header row found in {filename}")
        return None

    # Find other columns in the same header row
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or '').lower().replace('\n', ' ').replace('_x000d_', ' ')
        if 'number' in val and 'free' in val and 'reduced' not in val:
            col_free = c
        elif 'number' in val and 'reduced' in val:
            col_reduced = c
        elif 'number' in val and 'paid' in val:
            col_paid = c

    if col_enrollment is None or col_free is None:
        print(f"    Missing required columns in {filename}")
        return None

    # Aggregate school-level data
    total_enrollment = 0
    total_free = 0
    total_reduced = 0
    total_paid = 0
    school_count = 0

    for r in range(header_row + 1, ws.max_row + 1):
        enroll = ws.cell(row=r, column=col_enrollment).value
        if not isinstance(enroll, (int, float)) or enroll <= 0:
            continue

        free = ws.cell(row=r, column=col_free).value or 0
        reduced = ws.cell(row=r, column=col_reduced).value or 0 if col_reduced else 0
        paid = ws.cell(row=r, column=col_paid).value or 0 if col_paid else 0

        if not isinstance(free, (int, float)):
            free = 0
        if not isinstance(reduced, (int, float)):
            reduced = 0
        if not isinstance(paid, (int, float)):
            paid = 0

        total_enrollment += int(enroll)
        total_free += int(free)
        total_reduced += int(reduced)
        total_paid += int(paid)
        school_count += 1

    if total_enrollment == 0:
        return None

    frl_total = total_free + total_reduced
    pct_free = round(total_free / total_enrollment * 100, 1)
    pct_reduced = round(total_reduced / total_enrollment * 100, 1)
    pct_frl = round(frl_total / total_enrollment * 100, 1)

    return {
        'school_year': school_year,
        'grad_year': end_year,
        'total_enrollment': total_enrollment,
        'total_free': total_free,
        'total_reduced': total_reduced,
        'total_paid': total_paid,
        'pct_free': pct_free,
        'pct_reduced': pct_reduced,
        'pct_free_reduced_lunch': pct_frl,
        'school_count': school_count,
        'note': 'High FRL rates reflect Community Eligibility Provision (CEP), where entire schools qualify for free meals regardless of individual student eligibility.',
    }


def main():
    print("Processing KDE FRL qualifying data...")
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    results = {}
    for fname in sorted(os.listdir(RAW_DIR)):
        if not fname.endswith('.xlsx'):
            continue
        filepath = os.path.join(RAW_DIR, fname)
        print(f"  Processing {fname}...")
        data = extract_file(filepath)
        if data:
            sy = data['school_year']
            results[sy] = data
            print(f"    {sy}: {data['pct_free_reduced_lunch']}% FRL "
                  f"({data['total_enrollment']:,} students, {data['school_count']} schools)")

    out_path = os.path.join(PROCESSED_DIR, 'kde_frl_statewide.json')
    with open(out_path, 'w') as f:
        json.dump({
            'source': 'Kentucky Department of Education, Division of School and Community Nutrition, Qualifying Data Report',
            'by_school_year': results,
        }, f, indent=2)
    print(f"\n  Wrote {out_path}")
    print(f"  {len(results)} school years extracted")


if __name__ == '__main__':
    main()
