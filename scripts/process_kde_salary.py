"""Process KDE Average Classroom Teacher Salary Excel into JSON.

Reads:
  data/raw/kde_salary/Average Classroom Teacher Salaries (1989-2026).xlsx

Produces:
  data/processed/kde_salary_statewide.json   — statewide avg by school year
  data/processed/kde_salary_by_district.json  — district-level avg by school year + district code
"""

import json
import os
import re
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
RAW_PATH = os.path.join(DATA_DIR, 'raw', 'kde_salary',
                        'Average Classroom Teacher Salaries (1989-2026).xlsx')
PROCESSED_DIR = os.path.join(DATA_DIR, 'processed')


def parse_year_label(label):
    """Convert KDE year label like '2004-05' or '2024-25' to (start_year, end_year, school_year_str).
    Returns (2004, 2005, '2004-2005') for '2004-05'."""
    label = str(label).strip()
    m = re.match(r'(\d{4})-(\d{2,4})', label)
    if not m:
        return None
    start = int(m.group(1))
    end_part = m.group(2)
    if len(end_part) == 2:
        end = int(str(start)[:2] + end_part)
        if end <= start:
            end = int(str(start + 1)[:2] + end_part)
    else:
        end = int(end_part)
    school_year = f"{start}-{end}"
    return (start, end, school_year)


def extract_history_sheet(wb):
    """Extract statewide and district data from the history sheet."""
    ws = wb['History-Avg Classroom Teacher ']

    # Row 4 = headers
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]

    # Build year column map: col_index -> (start_year, end_year, school_year_str)
    year_cols = {}
    for i, h in enumerate(headers):
        if h and i >= 2:
            parsed = parse_year_label(h)
            if parsed:
                year_cols[i] = parsed

    # Find state average row (row 181 has "State Average by District salary" in col B)
    state_avg_row = None
    for row_idx in range(5, ws.max_row + 1):
        cell_b = str(ws.cell(row=row_idx, column=2).value or '').lower()
        if 'state average' in cell_b:
            state_avg_row = row_idx
            break

    # Extract statewide averages
    statewide = {}
    if state_avg_row:
        for col_idx, (start, end, sy) in year_cols.items():
            val = ws.cell(row=state_avg_row, column=col_idx + 1).value
            if val is not None:
                try:
                    statewide[sy] = round(float(val))
                except (ValueError, TypeError):
                    pass

    # Extract district data
    districts = {}
    for row_idx in range(5, ws.max_row + 1):
        dist_no = ws.cell(row=row_idx, column=1).value
        dist_name = ws.cell(row=row_idx, column=2).value
        if not dist_no or not dist_name:
            continue
        dist_no_str = str(dist_no).strip()
        dist_name_str = str(dist_name).strip()

        # Skip non-district rows
        if not re.match(r'^\d{3}$', dist_no_str):
            continue

        district_salaries = {}
        for col_idx, (start, end, sy) in year_cols.items():
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            if val is not None:
                try:
                    district_salaries[sy] = round(float(val))
                except (ValueError, TypeError):
                    pass

        if district_salaries:
            districts[dist_no_str] = {
                'name': dist_name_str.title().strip(),
                'kde_district_number': dist_no_str,
                'salaries_by_school_year': district_salaries,
            }

    return statewide, districts


def extract_current_year_sheet(wb):
    """Extract 2025-2026 data from the current year sheet."""
    ws = wb['2026 Average Classroom Teacher ']

    # Row 3 = header: "District", "2025-2026"
    year_label = ws.cell(row=3, column=2).value
    parsed = parse_year_label(year_label) if year_label else None
    if not parsed:
        return None, None, {}

    start, end, sy = parsed

    # Find state average (look for row with 'state average' that also has a numeric value in col B)
    state_avg = None
    for row_idx in range(5, ws.max_row + 1):
        cell_a = str(ws.cell(row=row_idx, column=1).value or '').lower()
        if 'state average' in cell_a:
            val = ws.cell(row=row_idx, column=2).value
            if val is not None:
                try:
                    state_avg = round(float(val))
                    break
                except (ValueError, TypeError):
                    continue

    # Extract district data
    districts = {}
    for row_idx in range(4, ws.max_row + 1):
        cell_a = str(ws.cell(row=row_idx, column=1).value or '').strip()
        m = re.match(r'^(\d{3})\s+(.+)$', cell_a)
        if not m:
            continue
        dist_no = m.group(1)
        dist_name = m.group(2).strip().title()
        val = ws.cell(row=row_idx, column=2).value
        if val is not None:
            try:
                districts[dist_no] = {
                    'name': dist_name,
                    'salary': round(float(val)),
                }
            except (ValueError, TypeError):
                pass

    return sy, state_avg, districts


def main():
    print("Processing KDE Average Classroom Teacher Salary data...")
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)

    # History: 1989-90 through 2024-25
    statewide_hist, districts_hist = extract_history_sheet(wb)
    print(f"  History sheet: {len(statewide_hist)} years of statewide data, {len(districts_hist)} districts")

    # Current year: 2025-26
    current_sy, current_state_avg, current_districts = extract_current_year_sheet(wb)
    print(f"  Current year sheet: {current_sy}, state avg ${current_state_avg:,}" if current_state_avg else "  No current year data")

    # Merge current year into statewide
    if current_sy and current_state_avg:
        statewide_hist[current_sy] = current_state_avg

    # Merge current year district data into history
    if current_sy and current_districts:
        for dist_no, info in current_districts.items():
            if dist_no in districts_hist:
                districts_hist[dist_no]['salaries_by_school_year'][current_sy] = info['salary']
            else:
                districts_hist[dist_no] = {
                    'name': info['name'],
                    'kde_district_number': dist_no,
                    'salaries_by_school_year': {current_sy: info['salary']},
                }

    # Save statewide
    statewide_out = {
        'source': 'Kentucky Department of Education, Office of Finance and Operations',
        'methodology': 'Average of district-level average classroom teacher salaries',
        'notes': 'Restricted to classroom teachers. Includes contract salary, extended days, and extra duty pay.',
        'by_school_year': statewide_hist,
    }
    path_sw = os.path.join(PROCESSED_DIR, 'kde_salary_statewide.json')
    with open(path_sw, 'w') as f:
        json.dump(statewide_out, f, indent=2)
    print(f"  Wrote {path_sw}")

    # Save district-level
    district_out = {
        'source': 'Kentucky Department of Education, Office of Finance and Operations',
        'methodology': 'Average classroom teacher salary per district',
        'districts': districts_hist,
    }
    path_dist = os.path.join(PROCESSED_DIR, 'kde_salary_by_district.json')
    with open(path_dist, 'w') as f:
        json.dump(district_out, f, indent=2)
    print(f"  Wrote {path_dist}")

    # Quick validation
    print("\n  Statewide salary spot check:")
    for sy in ['1989-1990', '1999-2000', '2004-2005', '2009-2010', '2019-2020', '2023-2024', '2024-2025', '2025-2026']:
        val = statewide_hist.get(sy)
        if val:
            print(f"    {sy}: ${val:,}")
        else:
            print(f"    {sy}: not found")

    print(f"\n  District count: {len(districts_hist)}")
    # Sample district
    if '001' in districts_hist:
        d = districts_hist['001']
        print(f"  Sample (001 {d['name']}): {len(d['salaries_by_school_year'])} years")
        for sy in ['1989-1990', '2024-2025']:
            v = d['salaries_by_school_year'].get(sy)
            if v:
                print(f"    {sy}: ${v:,}")


if __name__ == '__main__':
    main()
