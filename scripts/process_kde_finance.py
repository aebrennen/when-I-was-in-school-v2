"""Extract state-level per-pupil spending from KDE Annual Financial Reports.

Reads: data/raw/kde_finance/Revenues and Expenditures YYYY-YYYY.xlsx
Produces: data/processed/kde_finance_statewide.json
"""

import json
import os
import re
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
RAW_DIR = os.path.join(DATA_DIR, 'raw', 'kde_finance')
PROCESSED_DIR = os.path.join(DATA_DIR, 'processed')


def find_per_pupil_sheet(wb):
    """Find the expenditures per pupil sheet (name varies by year)."""
    for name in wb.sheetnames:
        if 'per pupil' in name.lower() or 'perpupil' in name.lower():
            return wb[name]
    return None


def find_state_per_pupil_row(ws):
    """Find the 'State Per Pupil' row."""
    for row_idx in range(ws.max_row, max(1, ws.max_row - 30), -1):
        for col in range(1, 4):
            val = str(ws.cell(row=row_idx, column=col).value or '').lower().strip()
            if 'state per pupil' in val or 'state per-pupil' in val:
                return row_idx
    return None


def find_column_by_header(ws, header_row, patterns):
    """Find column index matching any of the given patterns in the header row."""
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
        for pattern in patterns:
            if pattern in val:
                return col
    return None


def parse_numeric(val):
    """Parse a cell value that might be numeric or a formatted string like '$19,142'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val))
    s = str(val).replace('$', '').replace(',', '').strip()
    try:
        return round(float(s))
    except (ValueError, TypeError):
        return None


def extract_file(filepath):
    """Extract per-pupil spending from a single KDE finance file."""
    filename = os.path.basename(filepath)

    # Parse year from filename
    m = re.search(r'(\d{4})-(\d{4})', filename)
    if not m:
        return None
    start_year = int(m.group(1))
    end_year = int(m.group(2))
    school_year = f"{start_year}-{end_year}"
    fiscal_year = end_year  # FY aligns with spring

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = find_per_pupil_sheet(wb)
    if ws is None:
        print(f"    No per-pupil sheet found in {filename}")
        return None

    state_row = find_state_per_pupil_row(ws)
    if state_row is None:
        print(f"    No 'State Per Pupil' row found in {filename}")
        return None

    # Find header row (usually row 3 or 4)
    header_row = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').lower()
            if 'instruction' in val and '1000' in val:
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        # Try scanning more broadly
        for r in range(1, 10):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or '').lower()
                if 'total' in val and ('expense' in val or 'expend' in val):
                    header_row = r
                    break
            if header_row:
                break

    if header_row is None:
        print(f"    No header row found in {filename}")
        return None

    # Find key columns
    col_instruction = find_column_by_header(ws, header_row, ['instruction (1000)', 'instruction\n1000', 'instruction1000'])
    col_transport = find_column_by_header(ws, header_row, ['transportation (2700)', 'transportation\n2700', 'transp 2700', 'transportation2700', 'pupil transp'])
    col_facilities = find_column_by_header(ws, header_row, ['building improvement (4700)', 'bldg improve', 'facilities', 'building improvement4700'])
    col_total = find_column_by_header(ws, header_row, ['total expenses (1000-5100)', 'total expense', 'total expend'])

    # If we can't find total by name, try the last numeric column
    if col_total is None:
        for c in range(ws.max_column, 0, -1):
            val = str(ws.cell(row=header_row, column=c).value or '').lower()
            if 'total' in val:
                col_total = c
                break

    result = {
        'school_year': school_year,
        'fiscal_year': fiscal_year,
    }

    if col_total:
        result['per_pupil_total'] = parse_numeric(ws.cell(row=state_row, column=col_total).value)
    if col_instruction:
        result['per_pupil_instruction'] = parse_numeric(ws.cell(row=state_row, column=col_instruction).value)
    if col_transport:
        result['per_pupil_transportation'] = parse_numeric(ws.cell(row=state_row, column=col_transport).value)
    if col_facilities:
        result['per_pupil_facilities'] = parse_numeric(ws.cell(row=state_row, column=col_facilities).value)

    return result


def main():
    print("Processing KDE finance files...")
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    # Only process files for FY 2021+ (the gap years)
    target_files = []
    for fname in sorted(os.listdir(RAW_DIR)):
        if not fname.endswith('.xlsx'):
            continue
        m = re.search(r'(\d{4})-(\d{4})', fname)
        if m:
            end_year = int(m.group(2))
            if end_year >= 2021:  # FY 2021+
                target_files.append(os.path.join(RAW_DIR, fname))

    print(f"  Found {len(target_files)} files for FY 2021+")

    results = {}
    for filepath in target_files:
        fname = os.path.basename(filepath)
        print(f"  Processing {fname}...")
        data = extract_file(filepath)
        if data:
            fy = str(data['fiscal_year'])
            results[fy] = data
            pp = data.get('per_pupil_total')
            print(f"    FY {fy}: total=${pp:,}" if pp else f"    FY {fy}: no total found")

    out_path = os.path.join(PROCESSED_DIR, 'kde_finance_statewide.json')
    with open(out_path, 'w') as f:
        json.dump({
            'source': 'Kentucky Department of Education, Annual Financial Reports',
            'by_fiscal_year': results,
        }, f, indent=2)
    print(f"\n  Wrote {out_path}")
    print(f"  {len(results)} fiscal years extracted")

    # Spot check
    for fy in sorted(results.keys()):
        d = results[fy]
        print(f"    FY {fy} ({d['school_year']}): total=${d.get('per_pupil_total', 'N/A')}, "
              f"instr=${d.get('per_pupil_instruction', 'N/A')}, "
              f"transp=${d.get('per_pupil_transportation', 'N/A')}, "
              f"facil=${d.get('per_pupil_facilities', 'N/A')}")


if __name__ == '__main__':
    main()
